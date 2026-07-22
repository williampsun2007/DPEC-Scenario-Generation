'''
DPEC Scenario Generation Tool

Streamlit app that builds Breeze-ready emission control scenario files (.xlsx)
from DPEC pathway data.

Workflow:
1. User selects a base year and target year (which DPEC source workbooks to
   pull from), one or more DPEC scenarios, pollutant species, provinces,
   sectors, and weather (met) years via the UI.
2. For each selected species, every on/off combination is generated
   recursively (generate_scenarios), so N selected species yield up to 2^N
   variants per scenario.
3. For each variant, per-province/per-sector reduction percentages are
   computed as (base emissions - target emissions) / base emissions,
   pulled from the corresponding base-year and target-year source
   workbooks, and written into a copy of the Breeze template
   (管控方案模板.xlsx). Only province/sector pairs explicitly selected
   receive nonzero values; everything else stays at 0% change.
4. One output workbook is produced per (scenario x species-combination),
   duplicated once per selected weather year (data is identical across
   weather years -- only the filename changes) using the naming
   convention:
   {date}_{scenario}_{target_year}Target_{species_binary}_{base_year}Base_{weather_year}Met-{base_year}-{weather_year}
5. All generated workbooks are zipped in-memory and offered as a single
   download via the "Get Excel Files" / "Download all scenarios" buttons.

Expects source data files under data/ (per base year and per scenario/target
year) and a 管控方案模板.xlsx template in the working directory.
'''

import streamlit as st
import openpyxl
from datetime import date
import io
import zipfile
import time

if "base_year" not in st.session_state:
    st.session_state.base_year = "2017"
if "target_year" not in st.session_state:
    st.session_state.target_year = "2030"
if "scenarios" not in st.session_state:
    st.session_state.scenarios = "Baseline"
if "species" not in st.session_state:
    st.session_state.species = None
if "provinces" not in st.session_state:
    st.session_state.provinces = None
if "sectors" not in st.session_state:
    st.session_state.sectors = None
if "met_years" not in st.session_state:
    st.session_state.met_years = None
if "workbooks" not in st.session_state:
    st.session_state.workbooks = {}
    
province_cn_map = {
    "Beijing": "北京市",
    "Tianjin": "天津市",
    "Hebei": "河北省",
    "Shanxi": "山西省",
    "Inner Mongolia": "内蒙古自治区",
    "Liaoning": "辽宁省",
    "Jilin": "吉林省",
    "Heilongjiang": "黑龙江省",
    "Shanghai": "上海市",
    "Jiangsu": "江苏省",
    "Zhejiang": "浙江省",
    "Anhui": "安徽省",
    "Fujian": "福建省",
    "Jiangxi": "江西省",
    "Shandong": "山东省",
    "Henan": "河南省",
    "Hubei": "湖北省",
    "Hunan": "湖南省",
    "Guangdong": "广东省",
    "Guangxi": "广西壮族自治区",
    "Hainan": "海南省",
    "Chongqing": "重庆市",
    "Sichuan": "四川省",
    "Guizhou": "贵州省",
    "Yunnan": "云南省",
    "Tibet": "西藏自治区",
    "Shaanxi": "陕西省",
    "Gansu": "甘肃省",
    "Qinghai": "青海省",
    "Ningxia": "宁夏回族自治区",
    "Xinjiang": "新疆维吾尔自治区",
}

sector_cn_map = {
    "Power": "电力",
    "Industry": "工业",
    "Transportation": "交通",
    "Residential": "民用",
    "Agriculture": "农业",
}

st.title("DPEC Scenario Generation")

with st.expander("How this works", expanded = True):
    st.markdown("""
        This tool generates Breeze-ready emission control scenario files (`.xlsx`) from DPEC pathway data.

        **Selections**
        - **Base Year / Target Year** — pick one of each. Together these decide which DPEC source data is used to compute reduction percentages (target year emissions vs. base year emissions).
        - **Scenarios** - pick any subset
        - **Species** — pick any subset (SO₂, NOₓ, NH₃, VOC, PM). The tool automatically generates *every* possible on/off combination for just the species you selected — e.g. picking 3 species generates up to 8 files (each alone, every pair, all three together, and all off).
        - **Provinces / Sectors** — pick which ones should actually receive the real DPEC reduction values. Any province–sector pair outside your selection is left at 0% change, even for a species that's "on."
        - **Weather Years** — pick one or more. A separate copy of each scenario file is produced per weather year (the reduction data itself doesn't change across weather years — only the filename does, so Breeze knows which met year to pair it with).

        **What "Get Excel Files" produces**

        One `.xlsx` file per (scenario × species combination × weather year), all bundled into a single downloadable `.zip`.

        **Filename convention**

        `{date}_{scenario}_{target_year}Target_{species_binary}_{base_year}Base_{weather_year}Met-{base_year}-{weather_year}`

        - `date` — the day the batch was generated (YYYYMMDD)
        - `species_binary` — 5-digit code for which species are on (1) or off (0), in order SO₂-NOₓ-NH₃-VOC-PM

        *Example:* `20260709_EPNZCA_2035Target_11100_2020Base_2019Met-2020-2019` → EPNZCA scenario, 2035 target, SO₂+NOₓ+NH₃ on (VOC/PM off), 2020 baseline year, paired with 2019 weather data.
        
        **To avoid errors**, do not interact with any buttons/selections while excel files are generating.
    """)

st.session_state.base_year = st.pills("Base Year", ["2017", "2020"], selection_mode = "single",
                                      required = True, default = "2017")

target_year_options = ["2030", "2035", "2040", "2045", "2050", "2055", "2060"]
st.session_state.target_year = st.pills("Target Year", target_year_options, selection_mode = "single", 
                                        required = True, default = "2030")

scenario_options = ["Baseline", "CleanAir", "OTPCA", "OTPNZCA", "EPNZCA"]
st.session_state.scenarios = st.multiselect("Scenarios", scenario_options, default = [])

pollutant_options = ["SO2", "NOx", "NH3", "VOC", "PM25"]
st.session_state.species = st.multiselect("Species", pollutant_options, default = [])

province_list = [
    "Beijing", "Tianjin", "Hebei", "Shanxi", "Inner Mongolia",
    "Liaoning", "Jilin", "Heilongjiang", "Shanghai", "Jiangsu",
    "Zhejiang", "Anhui", "Fujian", "Jiangxi", "Shandong",
    "Henan", "Hubei", "Hunan", "Guangdong", "Guangxi",
    "Hainan", "Chongqing", "Sichuan", "Guizhou", "Yunnan",
    "Tibet", "Shaanxi", "Gansu", "Qinghai", "Ningxia",
    "Xinjiang"
]
st.session_state.provinces = st.multiselect("Provinces", province_list, default = [])
    
sector_list = ["Power", "Industry", "Transportation", "Residential", "Agriculture"]
st.session_state.sectors = st.multiselect("Sectors", sector_list, default = [])

met_years_list = list(range(2017, 2025))
st.session_state.met_years = st.multiselect("Weather Years", met_years_list, default = [])

def generate_scenarios(pos: int, bits: list, species: list, weather_years: list, scenario: str):
    if (pos < 5):
        bits[pos] = 0
        generate_scenarios(pos + 1, bits, species, weather_years, scenario)
    
        do_again = (pos == 0 and "SO2" in species) or (pos == 1 and "NOx" in species) \
                    or (pos == 2 and "NH3" in species) or (pos == 3 and "VOC" in species) \
                        or (pos == 4 and "PM25" in species)
    
        if (do_again):
            bits[pos] = 1
            generate_scenarios(pos + 1, bits, species, weather_years, scenario)
    else:
        wb_edit = openpyxl.load_workbook("管控方案模板.xlsx")
        ws_edit = wb_edit["管控方案"]
        ws_edit.delete_rows(3, 10000)
        for index, bit in enumerate(bits):
            if bit == 1:
                pollutant = pollutant_options[index]
                
                ws_base = wb_base[pollutant]
                ws_target = wb_target[pollutant]
                
                if (pollutant == "PM25"):
                    pollutant = "PM"
                elif (pollutant == "NOx"):
                    pollutant = "NOX"
            
                for province in st.session_state.provinces:
                    for sector in st.session_state.sectors:
                        row_base = sector_list.index(sector)
                        if (row_base == 2):
                            row_base = 3
                        elif (row_base == 3):
                            row_base = 2
                        
                        val_base = ws_base.cell(row = row_base + 2, column = province_list.index(province) + 2).value
                        val_target = ws_target.cell(row = sector_list.index(sector) + 3, column = province_list.index(province) + 3).value
                        
                        percentage = ((val_base - val_target) / val_base) * 100 if val_base != 0 else 0
                        if (percentage < 0):
                            percentage = 0
                            
                        ws_edit.append([province_cn_map[province], sector_cn_map[sector], pollutant, percentage])
              
        date_str = date.today().strftime("%Y%m%d")
        species_binary = ""
        for bit in bits:
            species_binary += str(bit)
                      
        for weather_year in weather_years:
            name = f'''{date_str}_{scenario}_{st.session_state.target_year}Target_{species_binary}_{st.session_state.base_year}Base_{weather_year}Met-{st.session_state.base_year}-{weather_year}.xlsx'''
            st.session_state.workbooks[name] = wb_edit  
            
num_excel_files = len(st.session_state.scenarios) * pow(2, len(st.session_state.species)) * len(st.session_state.met_years)
st.write(f"Number of Excel Files: {num_excel_files}")
                
if len(st.session_state.species) > 0 and len(st.session_state.provinces) > 0 and \
    len(st.session_state.sectors) > 0 and len(st.session_state.met_years) > 0 and len(st.session_state.scenarios) > 0:
        if st.button("Get Excel Files"):
            with st.spinner("Generating scenario files..."):
                time.sleep(2)
                st.session_state.workbooks = {}
                if st.session_state.base_year == "2017":
                    wb_base = openpyxl.load_workbook("data/2017_emission_report.xlsx")
                else:
                    wb_base = openpyxl.load_workbook("data/2020_emission_report.xlsx")
                    
                for scenario in st.session_state.scenarios:
                    wb_target = openpyxl.load_workbook(f"data/{scenario}_All_Years/{st.session_state.target_year}_Emission.xlsx")
            
                    generate_scenarios(0, [0] * 5, st.session_state.species, st.session_state.met_years, scenario)
            
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                    for filename, wb in st.session_state.workbooks.items():
                        excel_bytes = io.BytesIO()
                        wb.save(excel_bytes)
                        zf.writestr(filename, excel_bytes.getvalue())

                zip_buffer.seek(0)
                st.session_state.zip_bytes = zip_buffer.getvalue()

if "zip_bytes" in st.session_state:
    st.download_button("Download all scenarios (.zip)", st.session_state.zip_bytes, "dpec_scenarios.zip", "application/zip")
            
                        
                        
                        